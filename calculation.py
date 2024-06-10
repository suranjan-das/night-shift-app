# necessary import
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import timedelta
import pandas as pd
from copy import copy
import os

# prepare report file paths
gsheet_url = "https://docs.google.com/spreadsheets/d/1O2vnJiMRhnPZ0Z_5p8YbBlPxNkgGbcoppq8wU4iv_hU/export?format=csv"

def check_file_exists(file_path):
    """Checks if a file exists.

    Args:
        file_path: The path to the file.

    Raises:
        FileNotFoundError: If the file does not exist.
    """

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file {file_path} does not exist.")

def copy_range(source_wb: Workbook, destination_wb: Workbook, source_sheet_name: str,
               destination_sheet_name: str, row_range: tuple, source_col_num: int,
               destination_col_num: int, row_offset: int = 0) -> None:
    """
    Copy a range of data from one worksheet to another within Excel workbooks.

    Parameters:
        source_wb (Workbook): The source workbook object.
        destination_wb (Workbook): The destination workbook object.
        source_sheet_name (str): The name of the source worksheet.
        destination_sheet_name (str): The name of the destination worksheet.
        row_range (tuple): A tuple specifying the range of rows to copy (start, end, [step]).
        source_col_num (int): The column number (1-based index) in the source worksheet.
        destination_col_num (int): The column number (1-based index) in the destination worksheet.
        row_offset (int, optional): An optional row offset to apply when pasting data in the destination worksheet. Default is 0.

    Returns:
        None

    Example:
        copy_range(source_wb, destination_wb, 'Sheet1', 'Sheet2', (1, 10), 1, 2, row_offset=5)
    """
    # Load source worksheet
    source_ws: Worksheet = source_wb[source_sheet_name]

    # Load destination worksheet
    dest_ws: Worksheet = destination_wb[destination_sheet_name]

    # Iterate through rows in the source column
    for row_num in range(*row_range):
        # Copy and paste
        dest_ws.cell(row_num + row_offset, destination_col_num).value \
                = source_ws.cell(row_num, source_col_num).value

def copy_data(destination_wb, sheet_name, data, row_range, col_num):
    # Load destination workbook
    dest_ws = destination_wb[sheet_name]

    # Iterate through rows in the source column
    for i, row_num in enumerate(range(*row_range)):
        # copy and paste
        dest_ws.cell(row_num, col_num - 1).value = data[i]

def get_em_readings(url):
    night_df = pd.read_csv(url)
    night_df = night_df.fillna('')
    unit_readings = [night_df.iloc[i, 4] for i in range(2, 37)]
    line_readings = [night_df.iloc[i, 11] for i in range(2, 77)]
    return unit_readings, line_readings

def update_solar(source_wb, destination_wb, source_sheet_name, destination_sheet_name):
    # get solar file worksheets
    dgr_ws = source_wb[source_sheet_name]
    dest_ws = destination_wb[destination_sheet_name]
    # extract main dats
    solar_8MW_PTS_main = dgr_ws.cell(34, 4).value
    solar_8MW_TTS_main = dgr_ws.cell(34, 7).value
    solar_2MW_secA = dgr_ws.cell(34, 11).value
    solar_2MW_secB = dgr_ws.cell(34, 14).value
    # update apc file
    dest_ws.cell(40, 5).value = solar_8MW_PTS_main
    dest_ws.cell(42, 5).value = solar_8MW_TTS_main
    dest_ws.cell(44, 5).value = solar_2MW_secA
    dest_ws.cell(46, 5).value = solar_2MW_secB

def prepare_apc_sheet(date):
    # get report dates
    previous_report_date = date - timedelta(days=1)
    rdt = date.strftime('%d.%m.%Y')
    pdt = previous_report_date.strftime('%d.%m.%Y')
    # generate file paths
    final_apc_file_path = f"./generated/APC_{rdt}.xlsx"
    yesterday_apc_file_path = f"./uploads/APC_{pdt}.xlsx"
    pi_file_path = f"./uploads/PI {rdt}.xlsx"
    daily_gen_file_path = f"./uploads/DAILY GENERATION REPORT ON THE DATE {rdt}.xlsx"
    for path in [yesterday_apc_file_path, pi_file_path, daily_gen_file_path]:
        check_file_exists(path)
    # load the excel files into workbooks
    apc_previous = load_workbook(yesterday_apc_file_path)
    apc_today = copy(apc_previous)
    pi_today = load_workbook(pi_file_path)
    daily_gen_report = load_workbook(daily_gen_file_path)
    # copy previous day meter readings
    copy_range(apc_previous, apc_today, "Meter Reading",
            "Meter Reading", (4, 48), 5, 4)
    # copy previous day line readings
    copy_range(apc_previous, apc_today, "Meter Reading",
            "Meter Reading", (4, 79), 12, 11)
    # copy PI data
    copy_range(pi_today, apc_today, "Sheet1",
            "PI Data", (8, 81), 8, 4, row_offset=0)
    # copy solar data
    update_solar(daily_gen_report, apc_today, "DGR", "Meter Reading")
    # get todays reading from gsheet and update apc sheet
    unit_rdng, line_rdng = get_em_readings(gsheet_url)
    copy_data(apc_today, "Meter Reading", unit_rdng, (4, 38), 6)
    copy_data(apc_today, "Meter Reading", line_rdng, (4, 79), 13)
    # get GT-4 data and update (Pi data problem)
    apc_ws = apc_today["Meter Reading"]
    gen4 = (float(apc_ws.cell(11, 5).value)-float(apc_ws.cell(11, 4).value))*float(apc_ws.cell(11, 6).value)
    apc_today["PI Data"].cell(71, 4).value = str(round(gen4, 6))
    # some alignment adjustment
    currentCell = apc_ws.cell(73, 14)
    currentCell.alignment = Alignment(horizontal='right')
    # change apc todays date
    apc_today["PI Data"].cell(5, 2).value = rdt
    # save data
    apc_today.save(final_apc_file_path)
    apc_today.close()
    apc_previous.close()
    pi_today.close()
    daily_gen_report.close()

def prepare_sap_helper(date):
    daily_gen_file_path = f"./uploads/DAILY GENERATION REPORT ON THE DATE {date}.xlsx"
    daily_gen_report = load_workbook(daily_gen_file_path, data_only=True)
    helper_path = "./file_templates/SAP helper.xlsx"
    final_helper_path = "./generated/SAP helper.xlsx"
    check_file_exists(helper_path)
    sap_helper = load_workbook(helper_path)
    copy_range(daily_gen_report, sap_helper, "SAP POSTING DATA", "Sheet1", (2, 21), 3, 11)
    sap_helper.save(final_helper_path)
    sap_helper.close()
    daily_gen_report.close()