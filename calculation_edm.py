from openpyxl import load_workbook

import pandas as pd
import os
import re

def check_file_exists(file_path):
    """Checks if a file exists.

    Args:
        file_path: The path to the file.

    Raises:
        FileNotFoundError: If the file does not exist.
    """

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file {file_path} does not exist.")

def prepare_edm_file(date):
    # format date in different way
    rdt_1 = date.strftime('%d-%m-%Y')
    rdt_2 = date.strftime('%d_%m_%Y')
    rdt_3 = date.strftime('%d.%m.%Y')

    # prepare report file paths
    edm_file_path = "./file_templates/Daily EDM profiles Blockwise ABT Data upload template.xlsx"
    final_edm_file_path = "./generated/Daily EDM profiles Blockwise ABT Data upload.xlsx"
    injection_1_path = f"./uploads/FullSchedule-InjectionSummary-TSTPP-I-{rdt_1}.xlsx"
    injection_2_path = f"./uploads/FullSchedule-InjectionSummary-TALST2-{rdt_1}.xlsx"
    agc_1_path = f"./uploads/Talcher1_{rdt_2}.csv"
    agc_2_path = f"./uploads/Talcher2_{rdt_2}.csv"
    report_1_path = "./uploads/Report_1.xls"
    report_2_path = "./uploads/Report_2.xls"
    # pattern for injection profile files
    injection_1_pattern = f"FullSchedule-InjectionSummary-TSTPP-I\(\d+\)-{rdt_1}.xlsx"
    injection_2_pattern = f"FullSchedule-InjectionSummary-TALST2\(\d+\)-{rdt_1}.xlsx"
    # get the correct file name
    for file in os.listdir("./uploads"):
        if re.match(injection_1_pattern, file):
            injection_1_path = os.path.join("./uploads", file)
        if re.match(injection_2_pattern, file):
            injection_2_path = os.path.join("./uploads", file)
    # check if all required file exists
    paths = [edm_file_path, injection_1_path, injection_2_path,
             agc_1_path, agc_2_path, report_1_path, report_2_path]    
    for path in paths:
        check_file_exists(path)

    # read AG and DC data
    abt_1 = pd.read_html(report_1_path)[0]
    abt_2 = pd.read_html(report_2_path)[0]

    # read injection summary
    injection_1 = load_workbook(injection_1_path)
    injection_2 = load_workbook(injection_2_path)
    # load agc data
    agc_1 = pd.read_csv(agc_1_path)
    agc_2 = pd.read_csv(agc_2_path)
    agc_1.fillna(0, inplace=True)
    agc_2.fillna(0, inplace=True)
    
    # read EDM file
    edm = load_workbook(edm_file_path)
    stg1_ws = edm["Daily-Stg1"]
    stg2_ws = edm["Daily Stg-2"]

    # change date
    stg1_ws['AW5'].value = rdt_3
    stg2_ws['AW5'].value = rdt_3
    # make dc,ag and agc update
    for i in range(96):
        # DC update
        stg1_ws.cell(i+5, 2).value = abt_1["DC_MW"].iloc[i]
        stg2_ws.cell(i+5, 2).value = abt_2["DC_MW"].iloc[i]
        # AG update
        stg1_ws.cell(i+5, 5).value = abt_1["AG_MW"].iloc[i]
        stg2_ws.cell(i+5, 5).value = abt_2["AG_MW"].iloc[i]
        # AGC update
        stg1_ws.cell(i+5, 22).value = agc_1["AGC"].iloc[i]
        stg2_ws.cell(i+5, 22).value = agc_2["AGC"].iloc[i]

    inj1_ws = injection_1["Ex PP Summary "]
    inj2_ws = injection_2["Ex PP Summary "]
    # make injection profile update
    for c in range(15):
        for r in range(96):
            stg1_ws.cell(r+5, c+6).value = inj1_ws.cell(r+7, c+3).value
            stg2_ws.cell(r+5, c+6).value = inj2_ws.cell(r+7, c+3).value
    # save edm file
    edm.save(final_edm_file_path)
    edm.close()

# python rounding is funny, implemented excel rounding function
def excel_round(number, digits=2):
    s = f"{number:0.{digits+2}f}"[:-1]
    if int(s[-1]) >= 5:
        s = s[:-2] + str(int(s[-2]) + 1)
    else:
        s = s[:-1]
    return float(s)

def copy_excel_to_txt(excel_file, txt_file, sheet_name, date):
    rdt_3 = date.strftime('%d.%m.%Y')
    # Load the Excel workbook
    wb = load_workbook(excel_file, data_only=True)
    # Select the desired sheet (optional
    sheet = wb[sheet_name]

    with open(txt_file, 'w') as f:
        for row in range(22):
            s = []
            for col in range(99):
                if col == 0:
                    s.append(str(sheet['AV' + str(row+5)].value))
                elif col == 1:
                    s.append(rdt_3)
                elif col == 2:
                    s.append(sheet['AX' + str(row+5)].value)
                else:
                    if row == 0:
                        val = sheet['B' + str(col + 2)].value
                        s.append(f"{excel_round(val, 2):0.2f}")
                    elif row == 1:
                        val = sheet['D' + str(col + 2)].value
                        if val is not None:
                            s.append(f"{excel_round(val, 2):0.2f}")
                        else:
                            s.append("")
                    elif row == 2:
                        val1 = sheet['B' + str(col + 2)].value
                        val2 = sheet['D' + str(col + 2)].value
                        if val2 is not None:
                            s.append(f"{excel_round(val1+val2,2):0.2f}")
                        else:
                            s.append(f"{excel_round(val1,2):0.2f}")
                    elif row == 3:
                        val = sheet['E' + str(col + 2)].value
                        s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 4:
                        val = sheet['F' + str(col + 2)].value
                        s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 5:
                        val = sheet['G' + str(col + 2)].value
                        if val is None or val == 0:
                            s.append("")
                        else:
                            s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 6:
                        val = sheet['H' + str(col + 2)].value
                        if val is None or val == 0:
                            s.append("")
                        else:
                            s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 7:
                        val = sheet['I' + str(col + 2)].value
                        if val is None or val == 0:
                            s.append("")
                        else:
                            s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 8:
                        val = sheet['L' + str(col + 2)].value
                        if val is None or val == 0:
                            s.append("")
                        else:
                            s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 9:
                        val = sheet['M' + str(col + 2)].value
                        if val <= 0:
                            s.append("")
                        else:
                            s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 10:
                        val = sheet['M' + str(col + 2)].value
                        if val >= 0:
                            s.append("")
                        else:
                            s.append(f"{-excel_round(0.0025 * val,5):0.5f}")
                    elif row == 11:
                        val = sheet['N' + str(col + 2)].value
                        if val <= 0:
                            s.append("")
                        else:
                            s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 12:
                        val = sheet['N' + str(col + 2)].value
                        if val >= 0:
                            s.append("")
                        else:
                            s.append(f"{-excel_round(0.0025 * val,5):0.5f}")
                    elif row == 13:
                        val = sheet['V' + str(col + 2)].value
                        if val <= 0:
                            s.append("")
                        else:
                            s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 14:
                        val = sheet['V' + str(col + 2)].value
                        if val >= 0:
                            s.append("")
                        else:
                            s.append(f"{-excel_round(0.0025 * val,5):0.5f}")
                    elif row == 15:
                        val = sheet['J' + str(col + 2)].value
                        if val is None or val == 0:
                            s.append("")
                        else:
                            s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 16:
                        val = 0.0
                        s.append(f"{val:0.5f}")
                    elif row == 17:
                        val = sheet['P' + str(col + 2)].value
                        if val is None or val == 0:
                            s.append("")
                        else:
                            s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 18:
                        val = 0.0
                        s.append(f"{val:0.5f}")
                    elif row == 19:
                        val = sheet['R' + str(col + 2)].value
                        if val is None or val == 0:
                            s.append("")
                        else:
                            s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 20:
                        val = sheet['S' + str(col + 2)].value
                        if val is None or val == 0:
                            s.append("")
                        else:
                            s.append(f"{excel_round(0.0025 * val,5):0.5f}")
                    elif row == 21:
                        val = sheet['T' + str(col + 2)].value
                        if val is None or val == 0:
                            s.append("")
                        else:
                            s.append(f"{excel_round(0.0025 * val,5):0.5f}")
            f.write('\t'.join(s))
            f.write('\n')

def prepare_txt_file(date):
    final_edm_file_path = "./generated/Daily EDM profiles Blockwise ABT Data upload.xlsx"
    check_file_exists(final_edm_file_path)
    # generate text file
    copy_excel_to_txt(final_edm_file_path, './generated/st1.txt', 'Daily-Stg1', date)
    copy_excel_to_txt(final_edm_file_path, './generated/st2.txt', 'Daily Stg-2', date)

def prepare_daily_data_entry(date):
    rdt_3 = date.strftime('%d.%m.%Y')
    daily_data_path = "./file_templates/Daily data entry Format-TSTPS template.xlsx"
    check_file_exists(daily_data_path)
    daily_final_path = "./generated/Daily data entry Format-TSTPS.xlsx"
    pi_file_path = f"./uploads/PI {rdt_3}.xlsx"
    check_file_exists(pi_file_path)
    # Load the Excel workbook
    dl = load_workbook(daily_data_path)
    # Select the desired sheet (optional)
    dl_sheet = dl['TALCHER KANIHA']
    # load pi data sheet
    pi_today = load_workbook(pi_file_path)
    pi_sheet = pi_today['Sheet1']
    # change date
    dl_sheet['B9'] = rdt_3
    # update unit generation
    for i in range(6):
        dl_sheet['F' + str(9+i)].value = pi_sheet['H' + str(8 + 2*i)].value
    check_file_exists('./generated/st1.txt')
    for i in range(22):
        with open('./generated/st1.txt', 'r') as f:
            lines = f.readlines()
            for i, line in enumerate(lines):
                line = line.strip().replace(' ', '0')
                sum = 0
                for val in line.split('\t')[3:]:
                    try:
                        sum += float(val)
                    except ValueError:
                        pass
                if i < 3:
                    dl_sheet['F' + str(15+i)].value = round(sum/4000, 6)
                else:
                    dl_sheet['F' + str(15+i)].value = round(sum/10, 6)

    st2_daily_entry = []
    check_file_exists('./generated/st2.txt')
    with open('./generated/st2.txt', 'r') as f:
        lines = f.readlines()
        for i, line in enumerate(lines):
            line = line.strip().replace(' ', '0')
            sum = 0
            for val in line.split('\t')[3:]:
                try:
                    sum += float(val)
                except ValueError:
                    pass
            if i < 3:
                st2_daily_entry.append(round(sum/4000, 6))
            else:
                st2_daily_entry.append(round(sum/10, 6))

    dl_sheet['F37'].value = st2_daily_entry[2]
    dl_sheet['F38'].value = st2_daily_entry[0]
    dl_sheet['F39'].value = st2_daily_entry[1]
    dl_sheet['F40'].value = st2_daily_entry[3]
    dl_sheet['F41'].value = st2_daily_entry[4]
    dl_sheet['F42'].value = st2_daily_entry[5]
    dl_sheet['F43'].value = st2_daily_entry[6]
    dl_sheet['F44'].value = st2_daily_entry[7]
    dl_sheet['F45'].value = st2_daily_entry[8]
    dl_sheet['F46'].value = st2_daily_entry[11]
    dl_sheet['F47'].value = st2_daily_entry[12]
    dl_sheet['F48'].value = st2_daily_entry[13]
    dl_sheet['F49'].value = st2_daily_entry[14]
    dl_sheet['F50'].value = st2_daily_entry[15]
    dl_sheet['F51'].value = st2_daily_entry[16]
    dl_sheet['F52'].value = st2_daily_entry[9]
    dl_sheet['F53'].value = st2_daily_entry[10]
    dl_sheet['F54'].value = st2_daily_entry[17]
    dl_sheet['F55'].value = st2_daily_entry[18]
    dl_sheet['F56'].value = st2_daily_entry[19]
    dl_sheet['F57'].value = st2_daily_entry[20]
    dl_sheet['F58'].value = st2_daily_entry[21]

    dl.save(daily_final_path)
    dl.close()